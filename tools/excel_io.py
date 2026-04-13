"""Excel import/export for attendee lists and seat maps.

Pure functions — take data in, return data out. No DB access.
"""

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


# ── Column mapping: Excel header → attendee field ────────────
DEFAULT_COLUMN_MAP = {
    "姓名": "name",
    "name": "name",
    "职位": "title",
    "title": "title",
    "公司": "organization",
    "organization": "organization",
    "部门": "department",
    "department": "department",
    "角色": "role",
    "role": "role",
    "电话": "phone",
    "phone": "phone",
    "邮箱": "email",
    "email": "email",
}


def import_attendees_from_excel(
    file_path: Path | None = None,
    file_bytes: bytes | None = None,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Parse an Excel file into a list of attendee dicts.

    Args:
        file_path: Path to .xlsx file (mutually exclusive with file_bytes).
        file_bytes: Raw bytes of .xlsx file.
        column_map: Custom header→field mapping. Falls back to DEFAULT_COLUMN_MAP.

    Returns:
        List of attendee dicts with normalized field names.

    Raises:
        ValueError: If neither file_path nor file_bytes is provided,
                    or if no valid rows found.
    """
    if file_path is None and file_bytes is None:
        raise ValueError("Provide either file_path or file_bytes")

    cmap = {k.lower().strip(): v for k, v in (column_map or DEFAULT_COLUMN_MAP).items()}

    if file_bytes:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    else:
        wb = load_workbook(file_path, read_only=True)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("Excel file must have a header row and at least one data row")

    # Map headers
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    field_indices: dict[str, int] = {}
    for i, header in enumerate(headers):
        if header in cmap:
            field_indices[cmap[header]] = i

    if "name" not in field_indices:
        raise ValueError("Excel file must have a '姓名' or 'name' column")

    # Parse data rows
    attendees = []
    for row in rows[1:]:
        if not row or not any(row):
            continue
        att: dict[str, Any] = {"attrs": {}}
        for field, idx in field_indices.items():
            val = row[idx] if idx < len(row) else None
            att[field] = str(val).strip() if val is not None else None
        if att.get("name"):
            attendees.append(att)

    return attendees


def export_attendees_to_excel(
    attendees: list[dict[str, Any]],
    seats: list[dict[str, Any]] | None = None,
) -> bytes:
    """Export attendee list (optionally with seat info) to Excel bytes.

    Args:
        attendees: List of attendee dicts.
        seats: Optional list of seat dicts (matched by attendee_id).

    Returns:
        Bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "参会人员"

    # Build seat lookup
    seat_map: dict[str, dict] = {}
    if seats:
        for s in seats:
            if s.get("attendee_id"):
                seat_map[s["attendee_id"]] = s

    # Headers
    headers = ["姓名", "职位", "公司", "部门", "角色", "电话", "邮箱", "状态"]
    if seats:
        headers.extend(["座位号", "排", "列"])
    ws.append(headers)

    # Data rows
    for att in attendees:
        row = [
            att.get("name", ""),
            att.get("title", ""),
            att.get("organization", ""),
            att.get("department", ""),
            att.get("role", ""),
            att.get("phone", ""),
            att.get("email", ""),
            att.get("status", ""),
        ]
        if seats:
            seat = seat_map.get(att.get("id", ""), {})
            row.extend([
                seat.get("label", ""),
                seat.get("row_num", ""),
                seat.get("col_num", ""),
            ])
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_seatmap_to_excel(
    seats: list[dict[str, Any]],
    rows: int,
    cols: int,
) -> bytes:
    """Export a visual seat map grid to Excel.

    Each cell shows the attendee name or '空' for empty seats.

    Args:
        seats: List of seat dicts with row_num, col_num, and optional attendee_name.
        rows: Total rows in venue.
        cols: Total cols in venue.

    Returns:
        Bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "座位图"

    # Build lookup
    seat_map: dict[tuple[int, int], dict] = {}
    for s in seats:
        seat_map[(s["row_num"], s["col_num"])] = s

    # Header row (column numbers)
    ws.append([""] + [f"列{c}" for c in range(1, cols + 1)])

    # Data rows
    for r in range(1, rows + 1):
        row = [f"第{r}排"]
        for c in range(1, cols + 1):
            s = seat_map.get((r, c), {})
            if s.get("seat_type") in ("disabled", "aisle"):
                row.append("—")
            elif s.get("attendee_name"):
                row.append(s["attendee_name"])
            else:
                row.append("空")
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_excel_sheets_as_text(
    file_path: Path | None = None,
    file_bytes: bytes | None = None,
    *,
    max_rows_per_sheet: int = 30,
) -> str:
    """Read all sheets from an Excel file and format as plain text.

    Returns a human-readable summary suitable for LLM analysis.
    Each sheet is shown with its name, dimensions, and first N rows.
    """
    if file_path is None and file_bytes is None:
        raise ValueError("Provide either file_path or file_bytes")

    if file_bytes:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    else:
        wb = load_workbook(file_path, read_only=True)

    parts: list[str] = []
    for ws_sheet in wb.worksheets:
        title = ws_sheet.title or "Sheet"
        rows = list(ws_sheet.iter_rows(values_only=True))
        total = len(rows)
        parts.append(f"=== Sheet: {title} ({total} rows) ===")
        for i, row in enumerate(rows[:max_rows_per_sheet]):
            cells = [str(c) if c is not None else "" for c in row]
            # Trim trailing empty cells
            while cells and not cells[-1]:
                cells.pop()
            parts.append(f"  Row {i+1}: [{', '.join(cells)}]")
        if total > max_rows_per_sheet:
            parts.append(f"  ... ({total - max_rows_per_sheet} more rows)")
        parts.append("")

    wb.close()
    return "\n".join(parts)


def _is_label_cell(val: Any) -> bool:
    """Check if a cell value looks like a row/col label, not a person name."""
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    low = s.lower()
    # Row/col labels: "第一排", "A排", "Row 1", "第三列", column headers
    if any(kw in low for kw in [
        "排", "row", "列", "col", "座位", "编号", "号",
    ]):
        return True
    # Single letter (like "A", "B" used as row prefix)
    if len(s) == 1 and s.isalpha():
        return True
    return False


def _is_decoration_cell(val: Any) -> bool:
    """Check if a cell is a decoration element, not a name.

    Catches stage labels, aisle markers, backdrop labels, zone titles, etc.
    """
    if val is None:
        return True
    s = str(val).strip()
    if not s:
        return True
    # Normalize embedded newlines / whitespace for matching
    flat = "".join(s.split()).lower()
    # Stage, aisle, backdrop, and other decoration keywords
    if any(kw in flat for kw in [
        "舞", "台", "臺", "通道", "走道", "aisle", "stage",
        "背景", "墙", "牆", "门", "門", "入口", "出口",
        "讲台", "講臺", "主席台", "空", "—", "──",
        "座席", "席位", "区域", "區域",
        "贵宾", "貴賓", "嘉宾", "嘉賓", "观众", "觀眾",
        "vip", "reserved",
    ]):
        return True
    return False


def _is_name_cell(val: Any) -> bool:
    """Check if a cell value looks like a person name.

    Heuristic: non-empty, not a label, not a decoration,
    contains at least one CJK char or one alpha char.
    """
    if val is None:
        return False
    s = str(val).strip()
    if not s or len(s) > 30:
        return False
    if _is_label_cell(val) or _is_decoration_cell(val):
        return False
    has_cjk = any("\u4e00" <= c <= "\u9fff" for c in s)
    has_alpha = any(c.isalpha() for c in s)
    return has_cjk or has_alpha


def _count_seats_in_row(cells: tuple | list) -> int:
    """Count non-empty name cells in a row, skipping label columns."""
    if not cells:
        return 0
    return sum(1 for c in cells if _is_name_cell(c))


# ── Structured seat-layout parser ────────────────────────────

def parse_seat_layout_structured(
    file_path: Path | None = None,
    file_bytes: bytes | None = None,
) -> dict[str, Any]:
    """Parse an Excel seat-chart into a structured description.

    Designed for spatial seat-layout Excels where:
    - Each sheet = one area (观众席, 贵宾区, 贵宾室, …)
    - Cells contain attendee names at their physical positions
    - Row/column labels, stage markers, aisles are decoration

    Returns:
        {
          "areas": [
            {
              "name": "观众席",          # normalized simplified Chinese
              "role": "观众",             # inferred from sheet name
              "rows": 9,
              "cols": 11,
              "has_aisle": true,
              "aisle_after_col": 5,       # aisle between col 5 and 6
              "has_stage": true,
              "stage_position": "top",
              "attendees": [
                {"name": "都永海", "row": 1, "col": 1},
                {"name": "陳旭", "row": 1, "col": 2},
                ...
              ]
            },
            ...
          ],
          "total_attendees": 127,
          "total_seats": 152,
          "dedup_warnings": ["陳強 appears in 贵宾区 and 贵宾室"]
        }
    """
    from tools.chinese_norm import (
        clean_name,
        infer_role_from_area_name,
        normalize_zone,
    )

    if file_path is None and file_bytes is None:
        raise ValueError("Provide either file_path or file_bytes")

    if file_bytes:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    else:
        wb = load_workbook(file_path, read_only=True)

    areas: list[dict[str, Any]] = []
    all_names: dict[str, list[str]] = {}  # name → [area_names]

    for ws_sheet in wb.worksheets:
        raw_title = ws_sheet.title.strip() if ws_sheet.title else "Sheet"
        # Skip non-seat sheets
        if any(kw in raw_title.lower() for kw in [
            "说明", "readme", "注意", "instruction", "目录",
        ]):
            continue

        area_name = normalize_zone(raw_title) or raw_title
        role = infer_role_from_area_name(raw_title)

        rows_data = list(ws_sheet.iter_rows(values_only=True))
        if not rows_data:
            continue

        area = _parse_single_sheet(rows_data, area_name, role)
        if not area:
            continue

        # Track names for dedup
        for att in area["attendees"]:
            all_names.setdefault(att["name"], []).append(area_name)

        areas.append(area)

    wb.close()

    # Dedup warnings
    warnings = []
    for name, area_list in all_names.items():
        if len(area_list) > 1:
            locs = " 和 ".join(sorted(set(area_list)))
            warnings.append(f"{name} 同时出现在 {locs}")

    total_att = sum(len(a["attendees"]) for a in areas)
    total_seats = sum(a["rows"] * a["cols"] for a in areas)

    return {
        "areas": areas,
        "total_attendees": total_att,
        "total_seats": total_seats,
        "dedup_warnings": warnings,
    }


def _parse_single_sheet(
    rows_data: list[tuple],
    area_name: str,
    role: str,
) -> dict[str, Any] | None:
    """Parse one worksheet into an area dict with attendees + positions.

    Handles diverse layouts:
    - Header rows with column labels (第一列 ... 第十一列)
    - Row labels in leftmost column (第一排, 第二排, ...)
    - Empty rows between data rows (spacers)
    - Decoration cells (舞台, 通道, 背景牆)
    - Aisles (consecutive empty cols within a data row)
    - Irregular grids (贵宾室-style free-form seating)
    """
    from tools.chinese_norm import clean_name

    if not rows_data:
        return None

    # ── Phase 1: Identify data rows vs decoration ──────────
    # A "data row" has at least 1 name cell.
    # Collect all data rows with their original row index.
    data_rows: list[tuple[int, list]] = []  # (orig_idx, cells)
    stage_detected = False
    stage_position = "top"  # default

    for idx, raw_row in enumerate(rows_data):
        cells = list(raw_row) if raw_row else []
        # Trim trailing None
        while cells and cells[-1] is None:
            cells.pop()

        # Check for stage/decoration row
        joined = " ".join(
            str(c) for c in cells if c is not None
        ).strip()
        if any(kw in joined for kw in [
            "舞", "台", "臺", "stage", "背景", "牆", "墙",
        ]):
            stage_detected = True
            if data_rows:
                stage_position = "bottom"
            else:
                stage_position = "top"
            continue

        # Count name cells
        name_count = sum(1 for c in cells if _is_name_cell(c))
        if name_count >= 1:
            data_rows.append((idx, cells))

    if not data_rows:
        return None

    # ── Phase 2: Determine column bounds ────────────────────
    # Find the leftmost and rightmost name-bearing column indices
    # across all data rows.  Ignore label columns.
    all_col_indices: set[int] = set()
    for _, cells in data_rows:
        for ci, c in enumerate(cells):
            if _is_name_cell(c):
                all_col_indices.add(ci)

    if not all_col_indices:
        return None

    min_col = min(all_col_indices)
    max_col = max(all_col_indices)

    # ── Phase 3: Detect aisles (empty columns within data) ──
    # An aisle = a column index in [min_col, max_col] that has
    # zero names across ALL data rows.
    aisle_cols: set[int] = set()
    for ci in range(min_col, max_col + 1):
        if ci not in all_col_indices:
            # Verify it's truly empty in all rows
            has_name = False
            for _, cells in data_rows:
                if ci < len(cells) and _is_name_cell(cells[ci]):
                    has_name = True
                    break
            if not has_name:
                aisle_cols.add(ci)

    # Build col_index → logical_col mapping (skipping aisles)
    logical_col_map: dict[int, int] = {}
    logical_c = 1
    for ci in range(min_col, max_col + 1):
        if ci in aisle_cols:
            continue
        logical_col_map[ci] = logical_c
        logical_c += 1

    total_logical_cols = logical_c - 1

    # Detect primary aisle position (for metadata)
    has_aisle = len(aisle_cols) > 0
    aisle_after_col: int | None = None
    if aisle_cols:
        # Find the aisle gap: the logical col just before the aisle block
        sorted_aisles = sorted(aisle_cols)
        for ac in sorted_aisles:
            if ac - 1 >= min_col and ac - 1 in logical_col_map:
                aisle_after_col = logical_col_map[ac - 1]
                break

    # ── Phase 4: Extract attendees with positions ───────────
    attendees: list[dict[str, Any]] = []
    logical_row = 0

    for _, cells in data_rows:
        logical_row += 1
        for ci in range(min_col, max_col + 1):
            if ci in aisle_cols:
                continue
            if ci >= len(cells):
                continue
            val = cells[ci]
            if not _is_name_cell(val):
                continue
            name = clean_name(str(val).strip())
            if name:
                attendees.append({
                    "name": name,
                    "row": logical_row,
                    "col": logical_col_map[ci],
                })

    total_logical_rows = logical_row

    if not attendees:
        return None

    result: dict[str, Any] = {
        "name": area_name,
        "role": role,
        "rows": total_logical_rows,
        "cols": total_logical_cols,
        "has_aisle": has_aisle,
        "has_stage": stage_detected,
        "stage_position": stage_position,
        "attendees": attendees,
    }
    if aisle_after_col is not None:
        result["aisle_after_col"] = aisle_after_col

    return result


def parse_seat_layout_from_excel(
    file_path: Path | None = None,
    file_bytes: bytes | None = None,
) -> list[dict[str, Any]]:
    """Parse an Excel file into row_specs for generate_custom_layout.

    Each sheet = one zone (sheet name = zone label).
    Each data row = one seating row. Non-empty cells = seats.
    Header rows and noise rows (≤2 seats) are auto-filtered.

    Returns:
        List of row_spec dicts: [{count, repeat?, zone?}]
        compatible with generate_custom_layout().
    """
    # Delegate to structured parser, then flatten to row_specs
    structured = parse_seat_layout_structured(
        file_path=file_path, file_bytes=file_bytes,
    )

    all_specs: list[dict[str, Any]] = []
    for area in structured.get("areas", []):
        # Group attendees by row to get per-row seat counts
        rows_seats: dict[int, int] = {}
        for att in area["attendees"]:
            r = att["row"]
            rows_seats[r] = rows_seats.get(r, 0) + 1

        # Also account for empty seats (cols - occupied)
        max_cols = area.get("cols", 0)

        # Build row_specs grouped by consecutive equal counts
        sorted_rows = sorted(rows_seats.keys())
        if not sorted_rows:
            continue

        prev_count = -1
        repeat = 0
        zone = area.get("name")

        for r in sorted_rows:
            cnt = max(rows_seats[r], max_cols)  # use max_cols as row width
            if cnt == prev_count:
                repeat += 1
            else:
                if prev_count > 0:
                    spec: dict[str, Any] = {
                        "count": prev_count,
                        "repeat": repeat,
                    }
                    if zone:
                        spec["zone"] = zone
                    all_specs.append(spec)
                prev_count = cnt
                repeat = 1

        if prev_count > 0:
            spec = {"count": prev_count, "repeat": repeat}
            if zone:
                spec["zone"] = zone
            all_specs.append(spec)

    return all_specs
