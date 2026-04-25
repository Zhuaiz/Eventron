"""LangChain tools for the seating agent — wrapping service calls.

Each tool is a pure async function that the LLM can call via tool_calling.
Tools get service instances via closure from ``make_seating_tools()``.

Design principle: tools are *thin wrappers* around services. They emit
**facts**, not verdicts. The LLM decides WHEN and WHY to call them and
how to interpret the results. When a tool needs deep reasoning over
messy text (e.g. parsing 公司+职位 merged columns from a roster), it
delegates to an LLM call internally rather than encoding regex rules.
"""

from __future__ import annotations

import json
import uuid
from typing import Any

from langchain_core.tools import tool


def make_seating_tools(
    event_id: str,
    seat_svc: Any,
    event_svc: Any,
    attendee_svc: Any,
    llm_factory: Any | None = None,
) -> list:
    """Build LangChain tools with services bound via closure.

    Args:
        event_id: UUID string of the current event.
        seat_svc: SeatingService instance.
        event_svc: EventService instance.
        attendee_svc: AttendeeService instance.

    Returns:
        List of LangChain tool objects ready for ``llm.bind_tools()``.
    """
    eid = uuid.UUID(event_id)

    @tool
    async def get_event_info() -> str:
        """获取当前活动的基本信息（名称、布局类型、行列数、状态）。"""
        ev = await event_svc.get_event(eid)
        return json.dumps({
            "name": ev.name,
            "layout_type": ev.layout_type,
            "venue_rows": ev.venue_rows,
            "venue_cols": ev.venue_cols,
            "status": ev.status,
            "location": ev.location,
        }, ensure_ascii=False)

    @tool
    async def view_seats() -> str:
        """查看当前座位状态：总数、已分配、空闲、各分区详情。每次操作后都应调用此工具验证结果。"""
        seats = await seat_svc.get_seats(eid)
        if not seats:
            return "当前没有座位。需要先创建布局。"
        occupied = sum(1 for s in seats if s.attendee_id)
        total = len(seats)
        zones: dict[str, int] = {}
        for s in seats:
            z = getattr(s, "zone", None) or "未分区"
            zones[z] = zones.get(z, 0) + 1
        zone_str = ", ".join(
            f"{z}({c})" for z, c in sorted(zones.items())
        )
        max_row = max(s.row_num for s in seats)
        max_col = max(s.col_num for s in seats)
        return (
            f"总计 {total} 座, 已分配 {occupied}, "
            f"空闲 {total - occupied}。"
            f"布局 {max_row}排×{max_col}列。"
            f"分区: {zone_str}"
        )

    @tool
    async def create_layout(
        layout_type: str,
        rows: int,
        cols: int,
        table_size: int = 8,
        confirm_unusual: bool = False,
    ) -> str:
        """创建座位布局（替换已有座位）。

        如果 rows/cols 看起来"不太正常"（极端比例、远超人数、维度超大），
        工具会拒绝执行并请你用 confirm_unusual=True 重试。这不是硬限制，
        是给你一个反思机会：尤其要警惕"用 Excel 行数当场地排数"这种典型
        失误。当你**真的**知道场地很大或形状特殊时，再 confirm_unusual=True。

        Args:
            layout_type: 布局类型 grid|theater|roundtable|banquet|u_shape|classroom
            rows: 排数
            cols: 每排座位数
            table_size: 每桌人数（仅 roundtable/banquet 用）
            confirm_unusual: 已经核实过"奇怪"维度时设为 True，跳过反思阻拦
        """
        if rows is None or cols is None or rows < 1 or cols < 1:
            return "❌ rows/cols 必须是正整数。"

        # Soft sanity guards — only fire when something looks suspicious.
        # The agent can override with confirm_unusual=True after verifying.
        warnings: list[str] = []
        ratio = max(rows, cols) / max(1, min(rows, cols))
        if ratio > 6:
            warnings.append(
                f"维度比例 {rows}×{cols}（{ratio:.0f}:1）非常失衡，"
                "通常是把 Excel 行数当成场地排数引起的。"
            )
        # Capacity sanity vs current attendee count
        try:
            attendees_now = await attendee_svc.list_attendees_for_event(eid)
            n_att = sum(1 for a in attendees_now if a.status != "cancelled")
        except Exception:
            n_att = 0
        if n_att and rows * cols > max(40, n_att * 5):
            warnings.append(
                f"总座位 {rows*cols} 远超参会人数 {n_att}（>5×），"
                "可能是单维选大了。"
            )
        if max(rows, cols) > 60:
            warnings.append(
                f"单维 {max(rows, cols)} 超过 60，超出常见会场尺寸。"
            )

        if warnings and not confirm_unusual:
            tips = "\n  - " + "\n  - ".join(warnings)
            return (
                f"⚠️ {rows}×{cols} 看起来不太对：{tips}\n"
                "如果你已经核实过场地确实就这么大，再次调用并加上 "
                "confirm_unusual=True；否则请先：\n"
                "  · list_attendees 看真实人数\n"
                "  · 询问用户场地形状/尺寸\n"
                "  · 或者用 suggest_venue_dims 让我推荐"
            )

        seats = await seat_svc.create_venue_layout(
            eid,
            layout_type=layout_type,
            rows=rows,
            cols=cols,
            table_size=table_size,
            replace=True,
        )
        return (
            f"已创建 {layout_type} 布局: "
            f"{rows}排×{cols}列, 共 {len(seats)} 个座位"
        )

    @tool
    async def create_custom_layout(row_specs_json: str) -> str:
        """创建自定义布局（每排座位数可不同）。

        Args:
            row_specs_json: JSON 数组，例如:
                [{"count":8,"repeat":3,"zone":"贵宾区"},
                 {"count":20,"repeat":12}]
                count=每排座位数, repeat=连续相同的排数, zone=分区名(可选)
        """
        specs = json.loads(row_specs_json)
        seats = await seat_svc.create_custom_layout(
            eid, specs, replace=True,
        )
        total_rows = sum(s.get("repeat", 1) for s in specs)
        return (
            f"已创建自定义布局: {total_rows}排, "
            f"共 {len(seats)} 个座位"
        )

    @tool
    async def auto_assign(
        strategy: str = "priority_first",
    ) -> str:
        """自动排座：把参会者分配到空座位。

        Args:
            strategy: 排座策略 random|priority_first|by_department|by_zone
        """
        # Count unassigned attendees first for overflow reporting
        attendees_list = await attendee_svc.list_attendees_for_event(eid)
        all_seats = await seat_svc.get_seats(eid)
        seated_ids = {str(s.attendee_id) for s in all_seats if s.attendee_id}
        unassigned = [
            a for a in attendees_list
            if a.status in ("confirmed", "pending") and str(a.id) not in seated_ids
        ]
        available = [s for s in all_seats if not s.attendee_id and s.seat_type == "normal"]

        assignments = await seat_svc.auto_assign(
            eid, strategy=strategy,
        )
        if not assignments:
            return "没有需要分配的参会者（都已有座位，或没有参会者）"

        overflow = len(unassigned) - len(assignments)
        result = f"已分配 {len(assignments)} 位参会者, 策略: {strategy}"
        if overflow > 0:
            result += f"\n⚠️ 还有 {overflow} 人未分配座位（座位不足，共 {len(available)} 个空座位，{len(unassigned)} 人待分配）"
        return result

    @tool
    async def set_zone(
        row_start: int,
        row_end: int,
        zone_name: str,
    ) -> str:
        """给指定排的座位设置分区。

        Args:
            row_start: 起始排号（含），从 1 开始
            row_end: 结束排号（含）
            zone_name: 分区名，如 "贵宾区"、"嘉宾区"、"普通区"
        """
        seats = await seat_svc.get_seats(eid)
        target_ids = [
            s.id for s in seats
            if row_start <= s.row_num <= row_end
        ]
        if not target_ids:
            return f"第 {row_start}-{row_end} 排没有座位"
        count = await seat_svc.bulk_update_zone(target_ids, zone_name)
        return f"已将第 {row_start}-{row_end} 排共 {count} 个座位设为 {zone_name}"

    @tool
    async def set_zone_unzoned(zone_name: str) -> str:
        """把所有还没有分区的座位设为指定分区。

        Args:
            zone_name: 分区名
        """
        seats = await seat_svc.get_seats(eid)
        target_ids = [
            s.id for s in seats
            if getattr(s, "zone", None) is None
        ]
        if not target_ids:
            return "没有未分区的座位"
        count = await seat_svc.bulk_update_zone(target_ids, zone_name)
        return f"已将 {count} 个未分区座位设为 {zone_name}"

    @tool
    async def read_event_excel() -> str:
        """读取本活动最近上传的 Excel 文件内容（原始文本）。
        用于分析座位表、参会名单等。如果用户提到"文件"、"座位表"、"名单"，先调用此工具查看内容。"""
        from tools.event_files import find_latest_file_by_type
        from tools.excel_io import read_excel_sheets_as_text

        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "本活动没有上传过 Excel 文件"
        text = read_excel_sheets_as_text(file_path=entry["path"])
        fname = entry.get("filename", "Excel")
        return f"文件: {fname}\n\n{text}"

    @tool
    async def list_attendees() -> str:
        """查看当前活动的参会者名单。"""
        attendees = await attendee_svc.list_attendees_for_event(eid)
        if not attendees:
            return "当前没有参会者"
        lines = [f"共 {len(attendees)} 位参会者:"]
        for a in attendees[:30]:
            zone = a.role or "参会者"
            lines.append(f"  - {a.name} ({zone})")
        if len(attendees) > 30:
            lines.append(f"  ... 及另外 {len(attendees) - 30} 位")
        return "\n".join(lines)

    @tool
    async def import_attendees(attendees_json: str) -> str:
        """批量导入参会者（**纯 CRUD 工具，不做任何字段拆分/补全**）。

        你（LLM）负责把每条记录已经拆好。如果原始 Excel 把"公司+职位"
        合并在一格，调用前请用 smart_import_roster；或者你自己读完
        Excel 后逐行拆开后再传给本工具。

        Args:
            attendees_json: JSON 数组，每项必须有 name，可选字段：
                role, organization, title, department, priority
                例: [{"name":"张三","organization":"XX公司","title":"总经理",
                     "role":"贵宾","priority":80}]

        返回：新增/更新/跳过 数量；按 (name, organization) 去重。
        """
        from tools.chinese_norm import clean_name

        data = json.loads(attendees_json)
        existing = await attendee_svc.list_attendees_for_event(eid)
        existing_map: dict[str, Any] = {a.name: a for a in existing}

        created = 0
        updated = 0
        skipped = 0
        seen: set[str] = set()
        for item in data:
            if not isinstance(item, dict):
                continue
            raw_name = (item.get("name") or "").strip()
            name = clean_name(raw_name) or ""
            if not name:
                skipped += 1
                continue

            org_val = (item.get("organization") or "").strip() or None
            title_val = (item.get("title") or "").strip() or None

            fields: dict[str, Any] = {}
            if item.get("role"):
                fields["role"] = item["role"]
            if org_val:
                fields["organization"] = org_val
            if title_val:
                fields["title"] = title_val
            if item.get("department"):
                fields["department"] = item["department"]
            if "priority" in item:
                try:
                    fields["priority"] = int(item["priority"])
                except (TypeError, ValueError):
                    pass

            dup_key = f"{name}|{org_val or ''}"
            if dup_key in seen:
                skipped += 1
                continue
            seen.add(dup_key)

            if name in existing_map:
                if fields:
                    try:
                        await attendee_svc.update_attendee(
                            existing_map[name].id, **fields,
                        )
                        updated += 1
                    except Exception:
                        skipped += 1
                else:
                    skipped += 1
            else:
                try:
                    await attendee_svc.create_attendee(
                        eid, name=name, **fields,
                    )
                    created += 1
                except Exception:
                    skipped += 1

        parts = []
        if created:
            parts.append(f"新增 {created} 人")
        if updated:
            parts.append(f"更新 {updated} 人")
        if skipped:
            parts.append(f"跳过 {skipped} 人（重复/无效）")
        return f"导入完成: {', '.join(parts)}" if parts else "没有需要导入的数据"

    @tool
    async def inspect_excel() -> str:
        """检查活动 Excel 的**结构**，返回纯事实供你判断（不做分类、不做拆分）。

        每个 sheet 报告：
        - total_rows, max_width
        - header_row（第一行原文）
        - sample_rows（接下来 5 行原文）
        - stage_words（命中的舞台/通道关键词，空 list 表示没有）
        - name_cell_count（CJK/字母单元格的粗略计数）
        - numeric_only_columns（纯数字列索引，常是序号列）

        你看完这些事实后，自行判断：
        - 如果是"花名册"（列名是姓名/公司/职位等）→ 用 smart_import_roster
        - 如果是"空间座位图"（含 stage_words、cell 散布）→ analyze_seat_chart
        - 不确定时直接问用户
        """
        from tools.event_files import find_latest_file_by_type
        from tools.excel_io import inspect_excel_structure

        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "本活动没有上传过 Excel 文件"
        info = inspect_excel_structure(file_path=entry["path"])
        info["filename"] = entry.get("filename", "Excel")
        return json.dumps(info, ensure_ascii=False, indent=2)

    @tool
    async def smart_import_roster(
        column_mapping_hint: str = "",
        default_role: str = "参会者",
    ) -> str:
        """读取 Excel 花名册，**用 LLM 拆分混合字段**后批量导入。

        适用于："公司"和"职位"被合并写在一格"的常见场景，例如:
            "北方石油國際有限公司 總經理"
            "中信財務（國際）有限公司 公司總經理"
        不靠 regex / 后缀表 — 真正用 LLM 在工具内部一次性把所有行拆好，
        再用 priority/role 推断写入 DB。

        Args:
            column_mapping_hint: 可选，告诉拆分 LLM 哪一列是什么。
                例: "col0=序号, col1=公司+职位混合, col2=姓名"
                留空时 LLM 会自己看头部和示例行猜列含义。
            default_role: 找不到角色信息时的默认 role，默认"参会者"

        前置条件：先 inspect_excel 看清结构，再调本工具。
        """
        if llm_factory is None:
            return (
                "❌ 没有 LLM 工厂，无法运行 smart_import_roster。"
                "请改用 read_event_excel + import_attendees 的手动流程。"
            )

        from tools.chinese_norm import clean_name
        from tools.event_files import find_latest_file_by_type
        from tools.excel_io import inspect_excel_structure

        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "本活动没有上传过 Excel 文件"

        # Read raw rows (use openpyxl directly to keep all cells, no
        # heuristic field detection).
        from openpyxl import load_workbook
        wb = load_workbook(entry["path"], read_only=True, data_only=True)
        all_raw_rows: list[list[Any]] = []
        for ws_sheet in wb.worksheets:
            for row in ws_sheet.iter_rows(values_only=True):
                cells = list(row or ())
                while cells and cells[-1] is None:
                    cells.pop()
                if not cells:
                    continue
                all_raw_rows.append(cells)
        wb.close()

        if len(all_raw_rows) < 2:
            return "Excel 行数不足（至少需要表头 + 1 行数据）"

        structure = inspect_excel_structure(file_path=entry["path"])

        llm = llm_factory("smart")
        if llm is None:
            return "❌ LLM 服务不可用。"

        # Ask the LLM to emit a per-row JSON parse. We send the structure
        # facts + ALL data rows in one shot; the LLM returns a list of
        # {name, organization, title, role, priority} dicts.
        system = (
            "你是 Eventron 的花名册解析器。给你一份 Excel 的结构信息和"
            "所有数据行，请逐行解析为标准化参会者记录。\n\n"
            "规则：\n"
            "1. 先看 header_row 和 sample_rows 弄清每一列的含义。\n"
            "2. 如果某一列把'公司'+'职位'合并在一起（例如"
            " '北方石油國際有限公司 總經理'），按你对中文公司命名习惯的"
            "理解把它拆成 organization + title。'中信財務（國際）有限公司"
            " 公司總經理' 应拆为 org='中信財務（國際）有限公司',"
            " title='公司總經理'，不要被'公司總經理'里的'公司'误导。\n"
            "3. role 是角色标签（贵宾/嘉宾/演讲嘉宾/工作人员/参会者）。"
            "如果原数据没有明确角色，根据 title 判断：董事长/总裁/CEO 类"
            "→'贵宾'；总经理/合伙人/总监 → '嘉宾'；其他 → 默认 role。\n"
            "4. priority 0-100：董事长/总裁=90, 总经理=80, 副总=70,"
            " 经理/主管=50, 其他=30。\n"
            "5. 姓名清洗：去掉中间空格（'李 虎' → '李虎'）。\n"
            "6. 看到序号列（纯数字）就忽略它，不是数据。\n"
            "7. 严格输出 JSON 数组，不要任何 markdown 包裹、注释、前后语。\n"
            '   每项格式: {"name":"...", "organization":"...", "title":"...",'
            ' "role":"...", "priority": int}\n'
            "8. 解析不出 name 的行直接跳过（不要伪造）。"
        )
        # Send a compact representation
        payload = {
            "default_role": default_role,
            "column_hint": column_mapping_hint or None,
            "structure": structure,
            "rows": all_raw_rows,
        }
        try:
            resp = await llm.ainvoke([
                {"role": "system", "content": system},
                {
                    "role": "user",
                    "content": (
                        "请解析以下 Excel 行：\n"
                        + json.dumps(payload, ensure_ascii=False)
                    ),
                },
            ])
            raw = resp.content if hasattr(resp, "content") else str(resp)
            if isinstance(raw, list):
                raw = next(
                    (p.get("text") for p in raw if isinstance(p, dict) and p.get("type") == "text"),
                    "",
                )
        except Exception as exc:
            return f"❌ LLM 解析失败：{exc}"

        # Robust JSON extraction
        parsed: list[dict] = []
        text = (raw or "").strip()
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:]
            text = text.strip()
        try:
            parsed = json.loads(text)
        except Exception:
            # Try to find the first JSON array
            import re as _re
            m = _re.search(r"\[[\s\S]*\]", text)
            if m:
                try:
                    parsed = json.loads(m.group(0))
                except Exception:
                    parsed = []
        if not isinstance(parsed, list) or not parsed:
            return (
                "❌ LLM 没有返回有效 JSON，请把 Excel 用 read_event_excel "
                "看完后改用 import_attendees 手动构造。"
            )

        # Persist
        existing = await attendee_svc.list_attendees_for_event(eid)
        existing_map = {a.name: a for a in existing}
        created = updated = skipped = 0
        seen: set[str] = set()
        for item in parsed:
            if not isinstance(item, dict):
                continue
            name = clean_name((item.get("name") or "").strip()) or ""
            if not name:
                skipped += 1
                continue
            org = (item.get("organization") or "").strip() or None
            title = (item.get("title") or "").strip() or None
            role = (item.get("role") or default_role).strip()
            priority = item.get("priority")
            try:
                priority = int(priority) if priority is not None else None
            except Exception:
                priority = None

            key = f"{name}|{org or ''}"
            if key in seen:
                skipped += 1
                continue
            seen.add(key)

            fields: dict[str, Any] = {"role": role}
            if org:
                fields["organization"] = org
            if title:
                fields["title"] = title
            if priority is not None:
                fields["priority"] = priority

            try:
                if name in existing_map:
                    await attendee_svc.update_attendee(
                        existing_map[name].id, **fields,
                    )
                    updated += 1
                else:
                    await attendee_svc.create_attendee(
                        eid, name=name, **fields,
                    )
                    created += 1
            except Exception:
                skipped += 1

        return (
            f"smart_import_roster 完成: 新增 {created}, 更新 {updated},"
            f" 跳过 {skipped}, LLM 解析 {len(parsed)} 行"
        )

    @tool
    async def suggest_venue_dims(
        attendees_count: int,
        layout_type: str = "theater",
        user_hints: str = "",
    ) -> str:
        """让 LLM 给场地推荐 (rows, cols)，并解释理由。

        不返回单一答案 — 返回 2-3 个候选 + 推荐理由，给你（agent）空间
        和用户对话。如果用户告诉你具体场地形状/尺寸，请直接用，不要先调
        本工具。

        Args:
            attendees_count: 实际参会人数（必须！别猜）
            layout_type: theater|classroom|roundtable|banquet|u_shape|grid
            user_hints: 用户给过的形状/尺寸暗示，原样转给 LLM。
                例: '会场长 30 米宽 18 米' / '剧院式带通道'
        """
        if llm_factory is None:
            return "❌ 没有 LLM 工厂；请直接和用户确认 rows/cols。"
        if not attendees_count or attendees_count <= 0:
            return "❌ 必须给出真实参会人数。"

        llm = llm_factory("fast")
        if llm is None:
            return "❌ LLM 服务不可用。"

        system = (
            "你是会场布局顾问。给你人数+布局类型+用户暗示，给出 2-3 个"
            "候选 rows×cols 方案。\n"
            "约束：\n"
            "- 总座位数 ≥ 人数，且不超过 1.5×人数（避免大量空座）\n"
            "- 单维不超过 60，否则拆区域更合理\n"
            "- 比例符合该 layout_type 习惯（theater 偏宽，u_shape 偏深）\n"
            "- 如果用户给了具体尺寸，按尺寸精确算\n"
            "严格输出 JSON: "
            '{"options":[{"rows":int,"cols":int,"capacity":int,'
            '"rationale":"为什么这样"}], "recommendation_index": int,'
            '"questions_to_user":["可选追问"]}'
        )
        try:
            resp = await llm.ainvoke([
                {"role": "system", "content": system},
                {
                    "role": "user",
                    "content": json.dumps({
                        "attendees": attendees_count,
                        "layout_type": layout_type,
                        "hints": user_hints or "",
                    }, ensure_ascii=False),
                },
            ])
            txt = resp.content if hasattr(resp, "content") else str(resp)
            if isinstance(txt, list):
                txt = next(
                    (p.get("text") for p in txt if isinstance(p, dict) and p.get("type") == "text"),
                    "",
                )
            return str(txt).strip()
        except Exception as exc:
            return f"❌ LLM 推荐失败：{exc}"

    @tool
    async def list_attendees_with_seats() -> str:
        """查看参会者及其座位分配详情。比 list_attendees 更详细，包含座位标签和分区。
        适用于需要了解谁坐在哪里的场景（换座、调整前先查看）。"""
        attendees = await attendee_svc.list_attendees_for_event(eid)
        seats = await seat_svc.get_seats(eid)
        if not attendees:
            return "当前没有参会者"

        # Build seat lookup: attendee_id -> seat
        seat_map: dict[str, Any] = {}
        for s in seats:
            if s.attendee_id:
                seat_map[str(s.attendee_id)] = s

        lines = [f"共 {len(attendees)} 位参会者:"]
        for a in attendees[:50]:
            role = a.role or "参会者"
            s = seat_map.get(str(a.id))
            if s:
                zone_str = f", 分区:{s.zone}" if s.zone else ""
                lines.append(
                    f"  - {a.name} ({role}) → 座位 {s.label}"
                    f" (第{s.row_num}排第{s.col_num}列{zone_str})"
                )
            else:
                lines.append(f"  - {a.name} ({role}) → 未分配座位")
        if len(attendees) > 50:
            lines.append(f"  ... 及另外 {len(attendees) - 50} 位")
        return "\n".join(lines)

    @tool
    async def swap_two_attendees(name_a: str, name_b: str) -> str:
        """交换两位参会者的座位。通过姓名查找参会者和他们的座位，然后互换。

        Args:
            name_a: 第一位参会者的姓名
            name_b: 第二位参会者的姓名
        """
        attendees = await attendee_svc.list_attendees_for_event(eid)
        att_a = next((a for a in attendees if a.name == name_a), None)
        att_b = next((a for a in attendees if a.name == name_b), None)

        if not att_a:
            return f"未找到参会者: {name_a}"
        if not att_b:
            return f"未找到参会者: {name_b}"

        seats = await seat_svc.get_seats(eid)
        seat_a = next(
            (s for s in seats if s.attendee_id == att_a.id), None,
        )
        seat_b = next(
            (s for s in seats if s.attendee_id == att_b.id), None,
        )

        if not seat_a:
            return f"{name_a} 目前没有分配座位，无法换座"
        if not seat_b:
            return f"{name_b} 目前没有分配座位，无法换座"

        await seat_svc.swap_seats(seat_a.id, seat_b.id)

        return (
            f"已交换座位: {name_a} ({seat_a.label}) ↔ "
            f"{name_b} ({seat_b.label})"
        )

    @tool
    async def reassign_attendee_seat(
        name: str, target_seat_label: str,
    ) -> str:
        """将指定参会者移动到目标座位。如果参会者已有座位，先取消原座位再分配新座位。

        Args:
            name: 参会者姓名
            target_seat_label: 目标座位标签，如 "A3"、"B5"
        """
        attendees = await attendee_svc.list_attendees_for_event(eid)
        att = next((a for a in attendees if a.name == name), None)
        if not att:
            return f"未找到参会者: {name}"

        seats = await seat_svc.get_seats(eid)
        target = next(
            (s for s in seats if s.label == target_seat_label), None,
        )
        if not target:
            return (
                f"未找到座位: {target_seat_label}。"
                f"可用座位标签示例: "
                + ", ".join(s.label for s in seats[:5])
            )

        if target.attendee_id and target.attendee_id != att.id:
            occupant = next(
                (a for a in attendees
                 if a.id == target.attendee_id), None,
            )
            occ_name = occupant.name if occupant else "未知"
            return (
                f"座位 {target_seat_label} 已被 {occ_name} 占用。"
                "请先换座或取消该座位分配。"
            )

        if target.seat_type in ("disabled", "aisle"):
            return f"座位 {target_seat_label} 类型为 {target.seat_type}，不可分配"

        # Unassign from current seat if any
        current = next(
            (s for s in seats if s.attendee_id == att.id), None,
        )
        if current:
            await seat_svc.unassign_seat(current.id)

        # Assign to new seat
        await seat_svc.assign_seat(target.id, att.id)
        old_label = current.label if current else "无"
        return (
            f"已将 {name} 从座位 {old_label} "
            f"移至座位 {target_seat_label}"
        )

    @tool
    async def unassign_attendee(name: str) -> str:
        """取消指定参会者的座位分配，使其变为未入座状态。

        Args:
            name: 参会者姓名
        """
        attendees = await attendee_svc.list_attendees_for_event(eid)
        att = next((a for a in attendees if a.name == name), None)
        if not att:
            return f"未找到参会者: {name}"

        seats = await seat_svc.get_seats(eid)
        current = next(
            (s for s in seats if s.attendee_id == att.id), None,
        )
        if not current:
            return f"{name} 目前没有分配座位"

        await seat_svc.unassign_seat(current.id)
        return f"已取消 {name} 的座位 {current.label}"

    @tool
    async def delete_attendee_by_name(name: str) -> str:
        """删除一名参会者（按姓名查找）。会先解除其座位再删人，**不可恢复**。

        Args:
            name: 要删除的参会者姓名（精确匹配）
        """
        attendees = await attendee_svc.list_attendees_for_event(eid)
        att = next((a for a in attendees if a.name == name), None)
        if not att:
            return f"未找到参会者: {name}"
        seats = await seat_svc.get_seats(eid)
        current = next((s for s in seats if s.attendee_id == att.id), None)
        if current:
            await seat_svc.unassign_seat(current.id)
        ok = await attendee_svc.delete_attendee(att.id)
        return f"已删除参会者 {name}" if ok else f"删除 {name} 失败"

    @tool
    async def delete_all_attendees(confirm: bool = False) -> str:
        """**清空当前活动的所有参会者**（不可恢复）。座位布局保留，所有人变未分配。

        必须设置 confirm=True 才会真删 — 这是给你（LLM）一个反思缓冲：
        在用户明确说"删除所有"/"清空"/"重新导入"之类意图后再 confirm。

        Args:
            confirm: True 才执行删除；False（默认）只返回当前人数，不动数据
        """
        attendees = await attendee_svc.list_attendees_for_event(eid)
        n = len(attendees)
        if n == 0:
            return "当前活动没有参会者，无需删除"
        if not confirm:
            return (
                f"⚠️ 当前有 {n} 位参会者，未删除。"
                "如确认要清空，请再次调用并设 confirm=True。"
            )
        deleted = await attendee_svc.delete_all_for_event(eid)
        return f"✅ 已清空 {deleted} 位参会者；座位仍在但全部变为未分配"

    @tool
    async def regenerate_roster_from_excel(
        column_mapping_hint: str = "",
        default_role: str = "参会者",
        confirm: bool = False,
    ) -> str:
        """**重新生成参会者名单**：清空旧名单 → 用 LLM 解析 Excel → 重新导入。

        用户说"重新生成参会人"/"重新导入名单"/"清掉再导一遍"时调用本工具。
        原子操作：先清空，再重新解析当前活动最近一份 Excel（同
        smart_import_roster 的 LLM 拆字段流程）。

        Args:
            column_mapping_hint: 给解析 LLM 的列含义提示（同 smart_import_roster）
            default_role: 默认角色
            confirm: True 才执行；False 只汇报会做什么
        """
        if llm_factory is None:
            return "❌ 没有 LLM 工厂，无法运行重新生成。"
        from tools.event_files import find_latest_file_by_type
        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "❌ 本活动没有上传过 Excel 文件，无法重新生成"

        existing = await attendee_svc.list_attendees_for_event(eid)
        n_old = len(existing)
        if not confirm:
            return (
                f"⚠️ 即将清空 {n_old} 位旧参会者并从 "
                f"《{entry.get('filename', 'Excel')}》重新导入。"
                "如确认请再次调用并设 confirm=True。"
            )

        deleted = await attendee_svc.delete_all_for_event(eid)
        # Re-run the smart import flow with the same closure-bound tool
        sub_result = await smart_import_roster.ainvoke({
            "column_mapping_hint": column_mapping_hint,
            "default_role": default_role,
        })
        return (
            f"✅ 重新生成完成：清空旧 {deleted} 人 → 重新导入 → {sub_result}"
        )

    # ── Structured seat-chart import ─────────────────────────
    @tool
    async def analyze_seat_chart() -> str:
        """**仅当你已经用 inspect_excel 确认这是空间座位图时**才调本工具。

        空间座位图 = 单元格按物理位置摆放人名、通常含"舞台/通道"装饰、
        sheet 宽度 ≥6 列、不像普通的"姓名/公司/职位"列式表格。

        本工具只做 I/O 解析，不做分类。如果你把花名册（如3 列报名表）
        丢给本工具，会返回 0 个区域 — 那是你的判断错误，请回到
        inspect_excel 看清结构后改用 smart_import_roster。
        """
        from tools.event_files import find_latest_file_by_type
        from tools.excel_io import parse_seat_layout_structured

        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "本活动没有上传过 Excel 文件"

        result = parse_seat_layout_structured(file_path=entry["path"])

        if not result.get("areas"):
            return (
                "未识别出任何空间座位区域（areas=0）。常见原因：\n"
                "  · 这其实是花名册而不是座位图 → 改用 smart_import_roster\n"
                "  · sheet 行/列结构不规则 → 用 read_event_excel 先看原文"
            )

        parts = [f"📊 座位表分析结果 (文件: {entry.get('filename', 'Excel')})"]
        parts.append(
            f"总计 {result['total_attendees']} 位参会者, "
            f"约 {result['total_seats']} 个座位"
        )

        for area in result["areas"]:
            aisle_info = ""
            if area.get("has_aisle") and area.get("aisle_after_col"):
                aisle_info = f", 通道在第{area['aisle_after_col']}列后"
            stage_info = ""
            if area.get("has_stage"):
                pos = "前方" if area["stage_position"] == "top" else "后方"
                stage_info = f", 舞台在{pos}"
            parts.append(
                f"\n📍 {area['name']} ({area['role']}): "
                f"{area['rows']}排×{area['cols']}列, "
                f"{len(area['attendees'])}人"
                f"{aisle_info}{stage_info}"
            )

        if result["dedup_warnings"]:
            parts.append(
                f"\n⚠️ 重复出现: {len(result['dedup_warnings'])} 人"
                f"（同一人出现在多个区域，导入时会去重）"
            )

        parts.append(
            "\n💡 使用 import_from_seat_chart 可按此结构一键创建区域"
            "、导入参会者、生成布局并自动排座。"
        )
        return "\n".join(parts)

    @tool
    async def import_from_seat_chart(
        skip_areas: str = "",
    ) -> str:
        """根据座位表 Excel 一键完成：创建区域→导入参会者→生成布局→排座。

        自动执行完整流程（原子操作，任何一步失败则全部回滚）：
        1. 解析 Excel 中每个 sheet 为一个区域
        2. 创建区域（含行列数、舞台、布局信息）
        3. 导入所有参会者（角色从区域名推断，位置保留）
        4. 为每个区域生成座位布局
        5. 按 priority_first 策略自动排座

        Args:
            skip_areas: 要跳过的区域名（逗号分隔），如 "贵宾室"
                        贵宾室常与贵宾区人员重叠，可选择跳过。
                        留空 = 全部导入。
        """
        from tools.chinese_norm import normalize_role
        from tools.event_files import find_latest_file_by_type
        from tools.excel_io import parse_seat_layout_structured

        entry = find_latest_file_by_type(event_id, "excel")
        if not entry:
            return "本活动没有上传过 Excel 文件"

        result = parse_seat_layout_structured(file_path=entry["path"])
        if not result["areas"]:
            return (
                "未识别出任何空间座位区域。如果这是花名册，请改用 "
                "smart_import_roster。如果你确认是座位图但格式不规则，"
                "用 read_event_excel 看原文后手动 create_area + "
                "import_attendees。"
            )

        skip_set = {
            s.strip()
            for s in skip_areas.split(",")
            if s.strip()
        }

        # ── Wrap entire import in a SAVEPOINT for atomicity ──
        # If any step fails, the DB rolls back to before the import
        # started — no orphaned areas or partial attendee imports.
        try:
            async with seat_svc.begin_nested():
                report, total_created, total_updated, assigned_count = (
                    await _do_import(
                        result, skip_set, normalize_role,
                    )
                )
        except Exception as exc:
            return (
                f"❌ 导入失败，已回滚所有操作。\n"
                f"错误: {exc!s}\n"
                f"请检查 Excel 文件格式后重试。"
            )

        all_attendees = (
            await attendee_svc.list_attendees_for_event(eid)
        )
        seated_ids = set()
        for s in await seat_svc.get_seats(eid):
            if s.attendee_id:
                seated_ids.add(str(s.attendee_id))
        unseated = [
            a for a in all_attendees
            if str(a.id) not in seated_ids
            and a.status in ("confirmed", "pending")
        ]

        report.append(
            f"\n📊 导入完成: "
            f"{total_created} 新增, {total_updated} 更新, "
            f"{assigned_count} 人已排座"
        )
        if unseated:
            names = ", ".join(a.name for a in unseated[:10])
            extra = (
                f" 等共 {len(unseated)} 人"
                if len(unseated) > 10 else ""
            )
            report.append(
                f"⚠️ {len(unseated)} 人未分配座位（座位不足）: "
                f"{names}{extra}"
            )

        return "\n".join(report)

    async def _do_import(
        parse_result: dict,
        skip_set: set[str],
        normalize_role,
    ) -> tuple[list[str], int, int, int]:
        """Inner import logic — runs inside a SAVEPOINT."""
        report: list[str] = ["🚀 开始导入座位表..."]
        imported_names: set[str] = set()
        total_created = 0
        total_updated = 0

        # Delete existing areas + their seats to start fresh
        existing_areas = await seat_svc.list_areas(eid)
        for ea in existing_areas:
            await seat_svc.delete_area(ea.id)
        if existing_areas:
            report.append(
                f"  已清除 {len(existing_areas)} 个旧区域"
            )

        # Also clear orphan seats (area_id=None) left by earlier
        # create_layout calls — otherwise the canvas piles up multiple
        # overlapping layouts.
        all_seats = await seat_svc.get_seats(eid)
        orphans = [s for s in all_seats if s.area_id is None]
        if orphans:
            await seat_svc.clear_all_seats(eid)
            report.append(
                f"  已清理 {len(orphans)} 个未归区的旧座位"
            )

        y_cursor = 0.0
        area_spacing = 80.0

        for area_data in parse_result["areas"]:
            area_name = area_data["name"]
            if area_name in skip_set:
                report.append(f"  ⏭️ 跳过区域: {area_name}")
                continue

            rows = area_data["rows"]
            cols = area_data["cols"]
            role = area_data["role"]
            has_stage = area_data.get("has_stage", False)

            area = await seat_svc.create_area(
                eid,
                name=area_name,
                layout_type="theater" if rows > 2 else "grid",
                rows=rows,
                cols=cols,
                display_order=len(
                    await seat_svc.list_areas(eid),
                ),
                offset_x=0.0,
                offset_y=y_cursor,
                stage_label="舞台" if has_stage else None,
            )
            report.append(
                f"  ✅ 区域 {area_name}: {rows}排×{cols}列"
            )

            seats = await seat_svc.generate_area_layout(
                eid, area.id,
            )
            report.append(f"     → 生成 {len(seats)} 个座位")

            all_seats = await seat_svc.get_seats(eid)
            area_seat_ids = [
                s.id for s in all_seats if s.area_id == area.id
            ]
            if area_seat_ids:
                await seat_svc.bulk_update_zone(
                    area_seat_ids, area_name,
                )

            # Import attendees
            area_created = 0
            area_updated = 0
            existing_attendees = (
                await attendee_svc.list_attendees_for_event(eid)
            )
            existing_map = {
                a.name: a for a in existing_attendees
            }
            base_priority = 80 if "贵宾" in role else (
                60 if "嘉宾" in role else 30
            )

            for att_data in area_data["attendees"]:
                name = att_data["name"]
                if name in imported_names:
                    continue
                imported_names.add(name)

                norm_role = normalize_role(role)
                fields = {
                    "role": norm_role,
                    "priority": base_priority,
                }

                if name in existing_map:
                    await attendee_svc.update_attendee(
                        existing_map[name].id, **fields,
                    )
                    area_updated += 1
                else:
                    await attendee_svc.create_attendee(
                        eid, name=name, **fields,
                    )
                    area_created += 1

            total_created += area_created
            total_updated += area_updated
            if area_created or area_updated:
                report.append(
                    f"     → 参会者: 新增 {area_created}, "
                    f"更新 {area_updated}"
                )

            y_cursor += rows * 46.0 + area_spacing

        # Auto-assign
        assignments = await seat_svc.auto_assign(
            eid, strategy="by_zone",
        )
        assigned_count = len(assignments) if assignments else 0

        return report, total_created, total_updated, assigned_count

    # ── Area (venue zone) management tools ────────────────────
    @tool
    async def list_areas() -> str:
        """查看当前活动的所有区域（如贵宾区、观众席等）。
        每个区域有独立的布局类型、行列数和画布偏移。"""
        areas = await seat_svc.list_areas(eid)
        if not areas:
            return "当前没有区域。可以用 create_area 创建。"
        lines = [f"共 {len(areas)} 个区域:"]
        for a in areas:
            seats = await seat_svc.get_seats(eid)
            area_seats = [s for s in seats if s.area_id == a.id]
            lines.append(
                f"  - {a.name} ({a.layout_type} {a.rows}×{a.cols})"
                f" 座位{len(area_seats)}个"
                f"{' · 舞台:' + a.stage_label if a.stage_label else ''}"
                f" [偏移 x={a.offset_x}, y={a.offset_y}]"
            )
        return "\n".join(lines)

    @tool
    async def create_area(
        name: str,
        layout_type: str = "grid",
        rows: int = 5,
        cols: int = 10,
        offset_x: float = 0.0,
        offset_y: float = 0.0,
        stage_label: str = "",
    ) -> str:
        """创建一个新的区域（场馆分区）。创建后可用 generate_area_layout 生成该区域的座位。

        Args:
            name: 区域名称，如 "观众席"、"贵宾区"、"贵宾室"
            layout_type: 布局类型 grid|theater|roundtable|banquet|u_shape|classroom
            rows: 排数
            cols: 每排座位数
            offset_x: 在画布上的水平偏移（多区域时用于错开位置）
            offset_y: 在画布上的垂直偏移
            stage_label: 舞台/讲台标签（可选），如 "舞台"、"主席台"
        """
        existing = await seat_svc.list_areas(eid)
        display_order = len(existing)
        area = await seat_svc.create_area(
            eid,
            name=name,
            layout_type=layout_type,
            rows=rows,
            cols=cols,
            display_order=display_order,
            offset_x=offset_x,
            offset_y=offset_y,
            stage_label=stage_label or None,
        )
        return (
            f"已创建区域: {area.name} ({layout_type} {rows}×{cols})"
            f" 偏移({offset_x},{offset_y})"
            f"\n接下来调用 generate_area_layout 为该区域生成座位。"
        )

    @tool
    async def generate_area_layout(area_name: str) -> str:
        """为指定区域生成座位布局（替换该区域已有座位，不影响其他区域）。

        Args:
            area_name: 区域名称（需先用 create_area 创建）
        """
        areas = await seat_svc.list_areas(eid)
        area = next((a for a in areas if a.name == area_name), None)
        if not area:
            names = ", ".join(a.name for a in areas) if areas else "无"
            return f"未找到区域: {area_name}。当前区域: {names}"
        seats = await seat_svc.generate_area_layout(eid, area.id)
        return (
            f"已为 {area_name} 生成 {len(seats)} 个座位"
            f" ({area.layout_type} {area.rows}×{area.cols})"
        )

    @tool
    async def delete_area(area_name: str) -> str:
        """删除指定区域及其所有座位。

        Args:
            area_name: 区域名称
        """
        areas = await seat_svc.list_areas(eid)
        area = next((a for a in areas if a.name == area_name), None)
        if not area:
            return f"未找到区域: {area_name}"
        deleted = await seat_svc.delete_area(area.id)
        if deleted:
            return f"已删除区域 {area_name} 及其所有座位"
        return f"删除失败: {area_name}"

    return [
        get_event_info,
        view_seats,
        create_layout,
        create_custom_layout,
        auto_assign,
        set_zone,
        set_zone_unzoned,
        read_event_excel,
        inspect_excel,
        smart_import_roster,
        regenerate_roster_from_excel,
        suggest_venue_dims,
        analyze_seat_chart,
        import_from_seat_chart,
        list_attendees,
        import_attendees,
        list_attendees_with_seats,
        swap_two_attendees,
        reassign_attendee_seat,
        unassign_attendee,
        delete_attendee_by_name,
        delete_all_attendees,
        list_areas,
        create_area,
        generate_area_layout,
        delete_area,
    ]
